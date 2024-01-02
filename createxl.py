import openpyxl

# Read data from the .lst file
with open('pipe.lst', 'r') as file:
    lines = file.readlines()

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Process the first line separately as the title row
title_elements = lines[0].strip().split(',')
for col_num, title in enumerate(title_elements, start=1):
    sheet.cell(row=1, column=col_num, value=title)

# Process the remaining lines and add an additional column
for row_num, line in enumerate(lines[1:], start=2):
    # Split the line into elements
    elements = line.strip().split(',')

    # Write the elements to the sheet
    #sheet.cell(row=row_num, column=1, value=elements[0])  # The additional column
    for col_num, value in enumerate(elements[1:], start=2):
        sheet.cell(row=row_num, column=col_num, value=value)

# Save the workbook to a file
workbook.save('output.xlsx')

print("Excel sheet created successfully.")