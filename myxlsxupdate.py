from openpyxl import load_workbook


def open_and_update(name,row_index,args,result,pipelines):
    
    fname=name
    
    workbook= load_workbook(filename=fname)
    sheet=workbook.get_sheet_by_name('Sheet1')
    #filling args first
    le = len(args)
    sargs=[]
    for i in range(2,le+2):
        sargs=args[i-2]
        for j in range(1,len(sargs)+1):
            sheet.cell(row=i,column=j,value=sargs[j-1])
        sl = len(sargs)
        for k in range(sl+1,sl+2):
            sheet.cell(row=i,column=k,value=pipelines[i-2])
            sheet.cell(row=i,column=k+1,value=result[i-2])

    workbook.save(filename=fname)
    return 1

